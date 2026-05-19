import pandas as pd
import os
import re
from datetime import datetime, date, timedelta
from chinese_calendar import is_workday
import openpyxl

def extract_zone(text):
    if pd.isna(text):
        return None
    match = re.search(r'【(.*)】', str(text))
    if match:
        return re.sub(r'[【】]', '', match.group(1))
    return None

def is_workday_exceeded(date_series, days_limit):
    today = datetime.now().date()
    def count_bus_days(start_date):
        if pd.isna(start_date):
            return 0
        try:
            current = start_date.date()
            count = 0
            while current < today:
                current += timedelta(days=1)
                if is_workday(current):
                    count += 1
            return count
        except Exception:
            return None
    result = date_series.apply(count_bus_days)
    return result.notna() & (result > days_limit)

def is_calendar_exceeded(date_series, days_limit):
    today = datetime.now().date()
    def count_cal_days(start_date):
        if pd.isna(start_date):
            return 0
        try:
            return (today - start_date.date()).days
        except Exception:
            return 0
    return date_series.apply(count_cal_days) > days_limit

def _extract_hyperlink_map(source_file, col_name='需求名称'):
    """用 openpyxl 读取源文件，提取 HYPERLINK 公式中的 URL，返回 {原始行号: url} 映射。"""
    try:
        wb = openpyxl.load_workbook(source_file, data_only=False)
    except Exception:
        return {}
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    col_idx = None
    for i, h in enumerate(header, 1):
        if h == col_name:
            col_idx = i
            break
    if col_idx is None:
        wb.close()
        return {}
    url_map = {}
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        val = str(cell.value) if cell.value else ''
        if '=HYPERLINK(' in val:
            # 兼容 =HYPERLINK(...) 和 ==HYPERLINK(...) 两种格式
            start = val.index('=HYPERLINK(') + len('=HYPERLINK(')
            rest = val[start:]
            first_q = rest.index('"')
            second_q = rest.index('"', first_q + 1)
            url_map[row] = rest[first_q + 1:second_q]
    wb.close()
    return url_map


def process_logic(source_file="raw_data.xlsx", output_dir=None, timestamp=None):
    if not os.path.exists(source_file):
        print(f"错误: 找不到原始文件 '{source_file}'。")
        return None

    if not timestamp:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    
    if not output_dir:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        output_dir = os.path.join(base_dir, "result")
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    print(f"[*] 正在处理原始数据: {source_file}")
    df = pd.read_excel(source_file)

    # 提取 HYPERLINK 公式中的 URL，作为"需求地址"列
    url_map = _extract_hyperlink_map(source_file, '需求名称')
    if url_map:
        df['_原始行号'] = range(2, len(df) + 2)
        df['需求地址'] = df['_原始行号'].map(url_map)
        df = df.drop(columns=['_原始行号'])
        print(f"[*] 已提取 {len(url_map)} 条需求地址")

    # 预处理
    df['专区'] = pd.Series(dtype='object')
    df['登记日期'] = pd.to_datetime(df['登记日期'], errors='coerce')
    df = df.dropna(subset=['登记日期'])
    df = df[df['登记日期'] >= '2026-01-01']

    # 将"需求地址"列放到"需求名称"后面
    if '需求地址' in df.columns:
        cols = list(df.columns)
        cols.remove('需求地址')
        name_idx = cols.index('需求名称')
        cols.insert(name_idx + 1, '需求地址')
        df = df[cols]

    # 逻辑分析
    cond_review = (df['需求评审状态'].isin(['需求设计', '规划评审'])) | (df['开发状态'] == '等待任务分配')
    cond_not_online = df['开发状态'] != '已上线'
    cond1 = cond_review & cond_not_online & (is_workday_exceeded(df['登记日期'], 5) | is_calendar_exceeded(df['登记日期'], 7))
    cols1 = ['专区', '合同编号', '需求编号', '需求名称', '需求地址', '产品名称', '登记人员', '登记日期', '需求评审状态']
    cols1 = [c for c in cols1 if c in df.columns]
    res1 = df[cond1][cols1]

    cond2 = ((df['需求评审状态'] == '评审完成') & (df['需求实际状态'].isna() | (df['需求实际状态'] == '开发中')) & (df['开发状态'] == '开发中') & (is_workday_exceeded(df['登记日期'], 10)))
    cols2 = ['专区', '合同编号', '需求编号', '需求名称', '需求地址', '产品名称', '登记人员', '登记日期', '需求评审状态', '需求责任人', '开发状态', '开发工作量评审']
    cols2 = [c for c in cols2 if c in df.columns]
    res2 = df[cond2][cols2]

    exclude_dev = ['开发中', '已上线', '待任务分配', '作废', '终止', '设计评审', '已完成']
    exclude_actual = ['已上线', '作废', '暂停']
    cond3 = ((df['需求评审状态'] == '评审完成') & (~df['开发状态'].isin(exclude_dev)) & (~df['需求实际状态'].isin(exclude_actual)) & (is_workday_exceeded(df['登记日期'], 15)))
    cols3 = ['专区', '合同编号', '需求编号', '需求名称', '需求地址', '产品名称', '登记人员', '登记日期', '需求评审状态', '需求责任人', '开发状态', '开发工作量评审', '需求实际状态']
    cols3 = [c for c in cols3 if c in df.columns]
    res3 = df[cond3][cols3]

    output_filename = f"需求分析结果_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, output_filename)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        res1.to_excel(writer, sheet_name="1_评审超时", index=False)
        res2.to_excel(writer, sheet_name="2_开发超时", index=False)
        res3.to_excel(writer, sheet_name="3_上线超期", index=False)

    print(f"【分析完成】结果保存至: {output_path}")
    return output_path

if __name__ == "__main__":
    pass